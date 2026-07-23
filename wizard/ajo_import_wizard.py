# -*- coding: utf-8 -*-
import base64
import io
import logging

from odoo import _, api, fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger("MAZ")

try:
    import openpyxl
except ImportError:
    openpyxl = None

# Column letters (as used in the "Manufacturing Form" export) mapped to their
# 1-indexed spreadsheet column number, for the flat/normalized line table
# (columns V:AL) that repeats for every cut piece of every window.
COL = {
    'material_type': 22,   # V
    'ajo_name': 23,         # W
    'item_code': 24,        # X  (window / finished-good code, e.g. "1-A1-W2")
    'profile_code': 25,     # Y
    'color_code': 26,       # Z
    'length': 27,           # AA
    'height': 28,           # AB
    'qty': 29,              # AC
    'unit': 30,              # AD
    'angle': 31,             # AE
    'project_code': 32,      # AF
    'pm_name': 33,           # AG
    'profile_brand': 34,     # AH
    'glass_thick': 35,       # AI
    'floor': 36,             # AJ
    'block': 37,             # AK
    'handle': 38,            # AL
}

MATERIAL_TYPE_MAP = {
    'aluminum profile': 'aluminum',
    'glass': 'glass',
    'steel': 'steel',
    'accessories/alum-glass-metal': 'accessory',
    'acp': 'acp',
    'aluminum composite panel (acp)': 'acp',
}

# TechDesign 9.0 "Fiche de fabrication" text export: one job per block
# (delimited by a "Job:" line), each with a handful of tables. Column list
# and the material type each table maps to.
TD_SECTION_COLUMNS = {
    'Profiles': (['Article no.', 'Description', 'Color', 'Qty', 'Length', 'Cut'], 'aluminum'),
    'Additional profiles:': (['Article no.', 'Description', 'Color', 'Qty', 'Length', 'Cut'], 'aluminum'),
    'Length accessories:': (['Article no.', 'Description', 'Color', 'Qty', 'Length'], 'accessory'),
    'Fittings': (['Article no.', 'Description', 'Color', 'Qty'], 'accessory'),
    'Glass & panel': (['Description', 'Qty', 'Letter', 'Width', 'Height'], 'glass'),
}
TD_SECTION_HEADERS = set(TD_SECTION_COLUMNS) | {'Job:'}

# Maps the literal unit text found in the source file to the matching Odoo
# uom.uom external id, so imported lines carry the same unit as the source
# document instead of silently inheriting whatever uom the product defaults to.
UOM_XMLID_MAP = {
    'mm': 'uom.product_uom_millimeter',
    'cm': 'uom.product_uom_cm',
    'm': 'uom.product_uom_meter',
    'pcs': 'uom.product_uom_unit',
    'pc': 'uom.product_uom_unit',
    'unit': 'uom.product_uom_unit',
    'units': 'uom.product_uom_unit',
}


def _td_num(value):
    """Best-effort float parse of a TechDesign numeric cell (handles thousands separators)."""
    if value in (None, ''):
        return None
    try:
        return float(str(value).strip().replace(',', ''))
    except ValueError:
        return None


class AjoImportWizard(models.TransientModel):
    _name = 'ajo.import.wizard'
    _description = 'Import AJO Manufacturing Form (Excel / TechDesign export)'

    file = fields.Binary(string='Manufacturing Form (.xlsx or TechDesign .txt)', required=True)
    filename = fields.Char(string='Filename')
    warehouse_id = fields.Many2one(
        'stock.warehouse', string='Warehouse Override',
        help='Optional. Leave empty to auto-create/reuse one warehouse per project '
             '(matched by Project Code). Set this to force every AJO order created '
             'by this import onto the same warehouse instead.',
    )
    log = fields.Text(string='Import Log', readonly=True)

    def action_import(self):
        self.ensure_one()
        content = base64.b64decode(self.file)
        messages = []
        # Tracks AJO Orders created/reused earlier in *this* run, so multiple
        # window blocks for the same AJO number (normal within one file) keep
        # accumulating lines, while an AJO number that already existed before
        # this run started (a re-import of the same file) is skipped instead
        # of piling up duplicate lines. Threaded through as a plain argument
        # (not stored on self) since recordsets don't allow ad-hoc attributes.
        import_state = {'touched_order_ids': set(), 'skipped_order_names': []}

        if (self.filename or '').lower().endswith('.txt'):
            orders = self._import_techdesign_content(content, messages, import_state)
        else:
            if not openpyxl:
                raise UserError(_('The "openpyxl" python library is required to import Excel files.'))
            try:
                workbook = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
            except Exception as exc:
                raise UserError(_('Could not read the uploaded file as an Excel workbook: %s') % exc)
            orders = self._import_sheet(workbook.worksheets[0], messages, import_state)

        if not orders:
            raise UserError(_('No AJO window blocks were found in this file.'))

        self.log = '\n'.join(messages)

        action = {
            'name': _('Imported AJO Orders'),
            'type': 'ir.actions.act_window',
            'res_model': 'ajo_order',
            'context': self.env.context,
        }
        if len(orders) == 1:
            action.update({
                'view_mode': 'form',
                'views': [(False, 'form')],
                'res_id': orders[0].id,
            })
        else:
            action.update({
                'view_mode': 'list,form',
                'views': [(False, 'list'), (False, 'form')],
                'domain': [('id', 'in', orders.ids)],
            })

        if import_state['skipped_order_names']:
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('AJO Order(s) already imported'),
                    'message': _(
                        'These AJO numbers already existed and were NOT re-imported '
                        '(no duplicate lines were created): %s'
                    ) % ', '.join(sorted(set(import_state['skipped_order_names']))),
                    'type': 'warning',
                    'sticky': True,
                    'next': action,
                },
            }
        return action

    # ------------------------------------------------------------------
    # Shared helpers (master data get-or-create, used by both parsers)
    # ------------------------------------------------------------------

    def _get_or_create_order(self, ajo_name, project_ref, project_code, pm_name,
                              date_value, block_value, floor_value, messages, import_state):
        """Returns (order, importable). `importable` is False only when an AJO
        Order with this name already existed BEFORE this import run started
        (i.e. a re-import of an already-imported file) - callers must then
        skip its lines to avoid duplicates. Multiple blocks for the same AJO
        number within one run/file keep accumulating normally."""
        order = self.env['ajo_order'].search([('name', '=', ajo_name)], limit=1)
        if order:
            return order, order.id in import_state['touched_order_ids']

        pm_user = self.env['res.users']
        if pm_name:
            pm_user = self.env['res.users'].search([('name', '=', pm_name)], limit=1)
            if not pm_user:
                messages.append(_('PM "%s" not found, defaulted to %s.') % (pm_name, self.env.user.name))

        vals = {
            'name': ajo_name,
            'project_ref': project_ref or ajo_name,
            'project_code': project_code or '',
            'pm_id': pm_user.id if pm_user else self.env.user.id,
            'block': str(block_value) if block_value not in (None, '') else '',
            'floor': str(floor_value) if floor_value not in (None, '') else '',
        }
        # Leave 'warehouse_id' unset unless explicitly overridden: ajo_order.create()
        # auto-creates/reuses one warehouse per project (by Project Code).
        if self.warehouse_id:
            vals['warehouse_id'] = self.warehouse_id.id
        if hasattr(date_value, 'date'):
            vals['date'] = date_value.date()
        elif date_value:
            vals['date'] = date_value

        order = self.env['ajo_order'].create(vals)
        import_state['touched_order_ids'].add(order.id)
        messages.append(_('Created AJO Order %s.') % ajo_name)
        return order, True

    def _get_or_create_window_product(self, window_no, project_ref, messages, extra_label=''):
        if not window_no:
            return self.env['product.product']
        code = str(window_no)
        product = self.env['product.product'].search([('default_code', '=', code)], limit=1)
        if product:
            return product
        name = '%s - %s' % (code, project_ref or '')
        if extra_label:
            name = '%s (%s)' % (name, extra_label)
        template = self.env['product.template'].create({
            'name': name,
            'default_code': code,
            'type': 'consu',
            'sale_ok': False,
            'purchase_ok': False,
        })
        messages.append(_('Created window item %s.') % code)
        return template.product_variant_id

    def _get_or_create_material_product(self, material_type, profile_code, color_code,
                                         profile_brand, messages):
        profile_code = str(profile_code)
        color_code = str(color_code) if color_code else ''

        alum_profile = self.env['alum_profile'].search([('code', '=', profile_code)], limit=1)
        if not alum_profile:
            alum_profile = self.env['alum_profile'].create({
                'name': profile_code,
                'code': profile_code,
                'brand': profile_brand or '',
            })
            messages.append(_('Created profile master %s.') % profile_code)

        color = self.env['product_color']
        if color_code:
            color = self.env['product_color'].search([('code', '=', color_code)], limit=1)
            if not color:
                color = self.env['product_color'].create({
                    'name': color_code,
                    'code': color_code,
                })
                messages.append(_('Created color master %s.') % color_code)

        domain = [
            ('alum_profile', '=', alum_profile.id),
            ('material_type', '=', material_type),
            ('color_id', '=', color.id if color else False),
        ]
        template = self.env['product.template'].search(domain, limit=1)
        if template:
            return template.product_variant_id

        template = self.env['product.template'].create({
            # 'name' is left unset: product.template._compute_name derives it
            # from alum_profile + color_id (+ length, in mm, for aluminum).
            # 'is_storable' and Category/Sub Category are filled in by
            # product.template.create()'s own logic - no need to set them here.
            'alum_profile': alum_profile.id,
            'color_id': color.id if color else False,
            'material_type': material_type,
            'type': 'consu',
        })
        messages.append(_('Created material product %s.') % template.name)
        return template.product_variant_id

    def _get_or_create_angle(self, angle_label):
        if not angle_label:
            return self.env['angle']
        angle = self.env['angle'].search([('name', '=', str(angle_label))], limit=1)
        if not angle:
            angle = self.env['angle'].create({'name': str(angle_label)})
        return angle

    def _get_uom(self, unit_label):
        """Resolve the literal unit text from the source file to a uom.uom
        record, so the imported line keeps the same unit as the document
        instead of inheriting whatever uom the product happens to default to."""
        key = str(unit_label).strip().lower() if unit_label else ''
        xmlid = UOM_XMLID_MAP.get(key, 'uom.product_uom_unit')
        return (
            self.env.ref(xmlid, raise_if_not_found=False)
            or self.env.ref('uom.product_uom_unit')
        )

    def _create_order_line(self, order, item_product, product, height, qty, angle_label,
                            width=0.0, uom=None):
        angle = self._get_or_create_angle(angle_label)
        vals = {
            'order_id': order.id,
            'item_ref': item_product.id if item_product else False,
            'product_id': product.id,
            'width': width,
            'height': height,
            'qty': qty,
            'angle': angle.id if angle else False,
        }
        if uom:
            vals['product_uom_id'] = uom.id
        self.env['ajo_order_line'].create(vals)

    # ------------------------------------------------------------------
    # Excel "Manufacturing Form" parsing
    # ------------------------------------------------------------------

    def _import_sheet(self, sheet, messages, import_state):
        orders = self.env['ajo_order']
        row = 1
        max_row = sheet.max_row
        while row <= max_row:
            if sheet.cell(row=row, column=1).value == 'ALUMEC':
                order, next_row = self._import_block(sheet, row, messages, import_state)
                if order:
                    orders |= order
                row = next_row
            else:
                row += 1
        return orders

    def _import_block(self, sheet, start_row, messages, import_state):
        """Parse one window block starting at `start_row` (the 'ALUMEC' row).
        Returns (ajo_order record, row where the next block may start)."""
        get = lambda r, c: sheet.cell(row=r, column=c).value

        ajo_name = get(start_row + 11, 5)
        window_no = get(start_row + 14, 5)
        if not ajo_name:
            messages.append(_('Row %s: skipped block with no AJO no.') % start_row)
            return None, start_row + 1

        project_ref = get(start_row + 3, 5)
        project_code = get(start_row + 4, 5)
        pm_name = get(start_row + 5, 5)
        date_value = get(start_row + 8, 5)
        block_value = get(start_row + 12, 5)
        floor_value = get(start_row + 13, 5)

        order, importable = self._get_or_create_order(
            ajo_name, project_ref, project_code, pm_name,
            date_value, block_value, floor_value, messages, import_state,
        )
        if not importable:
            import_state['skipped_order_names'].append(ajo_name)
            messages.append(_(
                'AJO Order %s already exists - this window block was NOT re-imported '
                '(no duplicate lines were created).'
            ) % ajo_name)
            return order, start_row + 1

        item_product = self._get_or_create_window_product(window_no, project_ref, messages)

        # Scan forward for the flat line table (columns V:AL) until the next
        # block ('ALUMEC' in column A) or the end of the sheet.
        row = start_row + 1
        max_row = sheet.max_row
        while row <= max_row and get(row, 1) != 'ALUMEC':
            material_label = get(row, COL['material_type'])
            profile_code = get(row, COL['profile_code'])
            label_key = str(material_label).strip().lower() if material_label else ''
            if label_key in MATERIAL_TYPE_MAP and profile_code not in (None, 0):
                self._import_line(sheet, row, order, item_product, messages)
            row += 1

        return order, row

    def _import_line(self, sheet, row, order, item_product, messages):
        get = lambda c: sheet.cell(row=row, column=c).value

        material_label = (get(COL['material_type']) or '').strip().lower()
        material_type = MATERIAL_TYPE_MAP.get(material_label)
        if not material_type:
            messages.append(_('Row %s: unknown material type "%s", skipped.') % (row, material_label))
            return

        profile_code = get(COL['profile_code'])
        color_code = get(COL['color_code'])
        profile_brand = get(COL['profile_brand'])
        length_value = get(COL['length'])
        qty_value = get(COL['qty']) or 0
        angle_label = get(COL['angle'])
        unit_value = get(COL['unit'])

        product = self._get_or_create_material_product(
            material_type, profile_code, color_code, profile_brand, messages,
        )

        try:
            height = float(length_value)
        except (TypeError, ValueError):
            height = 0.0
            messages.append(_('Row %s: non-numeric length "%s", set to 0.') % (row, length_value))

        # Unit is only ever meaningfully variable for Accessories (pcs vs a
        # total length in mm); every other material type is a measurement in
        # mm regardless of what the sheet's Unit column literally says.
        uom = self._get_uom(unit_value) if material_type == 'accessory' else self._get_uom('mm')
        self._create_order_line(order, item_product, product, height, qty_value, angle_label, uom=uom)

    # ------------------------------------------------------------------
    # TechDesign 9.0 "Fiche de fabrication" text export parsing
    # ------------------------------------------------------------------

    def _decode_techdesign_text(self, content):
        for encoding in ('utf-16', 'utf-8-sig', 'utf-8', 'latin-1'):
            try:
                return content.decode(encoding)
            except (UnicodeDecodeError, LookupError):
                continue
        return content.decode('utf-8', errors='ignore')

    def _import_techdesign_content(self, content, messages, import_state):
        text = self._decode_techdesign_text(content)
        lines = [line.strip('\r\n') for line in text.split('\n')]
        n = len(lines)

        job_starts = [i for i, line in enumerate(lines) if line.strip() == 'Job:']
        orders = self.env['ajo_order']
        for i, start in enumerate(job_starts):
            end = job_starts[i + 1] if i + 1 < len(job_starts) else n
            order = self._import_techdesign_block(lines, start, end, messages, import_state)
            if order:
                orders |= order
        return orders

    def _td_value_after(self, lines, label, start, end):
        for idx in range(start, end):
            if lines[idx].strip() == label:
                j = idx + 1
                while j < end and not lines[j].strip():
                    j += 1
                return lines[j].strip() if j < end else None
        return None

    def _td_system_name(self, lines, start, end):
        """The descriptive hardware-system name printed right after the
        'Handle_Height' / '<value> mm' pair (e.g. 'GY 2 rails standard 3D')."""
        for idx in range(start, end):
            if lines[idx].strip() == 'Handle_Height':
                j = idx + 2  # skip past the "<value> mm" line
                while j < end and not lines[j].strip():
                    j += 1
                return lines[j].strip() if j < end else None
        return None

    def _import_techdesign_block(self, lines, start, end, messages, import_state):
        job_name = self._td_value_after(lines, 'Job:', start, end)
        item_code = self._td_value_after(lines, 'Item n°:', start, end)
        job_no = self._td_value_after(lines, 'Job n°:', start, end)
        system_name = self._td_system_name(lines, start, end)

        if not job_name:
            messages.append(_('Line %s: skipped TechDesign block with no Job name.') % start)
            return None

        # The 'Job:' text is the one field that repeats identically across
        # every block in the file, so it is used as the AJO grouping key.
        order, importable = self._get_or_create_order(
            ajo_name=job_name,
            project_ref=job_name,
            project_code=job_no or '',
            pm_name=None,
            date_value=fields.Date.context_today(self),
            block_value='',
            floor_value='',
            messages=messages,
            import_state=import_state,
        )
        if not importable:
            import_state['skipped_order_names'].append(job_name)
            messages.append(_(
                'AJO Order %s already exists - this block was NOT re-imported '
                '(no duplicate lines were created).'
            ) % job_name)
            return order

        item_product = self._get_or_create_window_product(
            item_code, job_name, messages, extra_label=item_code and job_no or '',
        )

        for header, (columns, material_type) in TD_SECTION_COLUMNS.items():
            self._import_techdesign_section(
                lines, start, end, header, columns, material_type,
                order, item_product, system_name, messages,
            )
        return order

    def _import_techdesign_section(self, lines, start, end, header, columns, material_type,
                                    order, item_product, profile_brand, messages):
        header_idx = None
        for idx in range(start, end):
            if lines[idx].strip() == header:
                header_idx = idx
                break
        if header_idx is None:
            return

        # Skip past the column-name header lines that follow the section title.
        j = header_idx + 1
        for col in columns:
            if j < end and lines[j].strip() == col:
                j += 1

        ncols = len(columns)
        qty_pos = columns.index('Qty')
        while j < end:
            line = lines[j].strip()
            if not line:
                j += 1
                continue
            if line in TD_SECTION_HEADERS:
                break
            row = [c.strip() for c in lines[j:j + ncols]]
            if len(row) < ncols or _td_num(row[qty_pos]) is None:
                # Not a data row (e.g. a glass-type spec line) - skip ahead one
                # line at a time until the grid realigns.
                j += 1
                continue
            values = dict(zip(columns, row))
            self._import_techdesign_row(values, material_type, order, item_product, profile_brand, messages)
            j += ncols

    def _import_techdesign_row(self, values, material_type, order, item_product, profile_brand, messages):
        if material_type == 'glass':
            profile_code = values.get('Description')
            color_code = None
            width = _td_num(values.get('Width')) or 0.0
            height = _td_num(values.get('Height')) or 0.0
            qty = _td_num(values.get('Qty')) or 0.0
            angle_label = None
        else:
            profile_code = values.get('Article no.')
            color_code = values.get('Color')
            width = 0.0
            height = _td_num(values.get('Length')) or 0.0
            qty = _td_num(values.get('Qty')) or 0.0
            angle_label = values.get('Cut')

        if not profile_code:
            return

        product = self._get_or_create_material_product(
            material_type, profile_code, color_code, profile_brand, messages,
        )
        # TechDesign has no per-row Unit column: Accessories are counted in
        # pieces ("pcs"), everything else (profiles, glass, ...) is a
        # measurement in "mm".
        uom = self._get_uom('pcs' if material_type == 'accessory' else 'mm')
        self._create_order_line(order, item_product, product, height, qty, angle_label, width=width, uom=uom)
